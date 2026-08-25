"""Renders stored businesses into an .xlsx, in memory.

Nothing is written to disk. Serverless hosting has no persistent filesystem, so
the workbook is built fresh on every download rather than kept as a file. The
consequence, recorded so it is not rediscovered: rows typed directly into a
downloaded sheet cannot be read back, because there is no stored file to read.
"""

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.models import Business

SHEET = "Businesses"

# Header text -> Business field. Order here is the column order in the sheet.
_LAYOUT: list[tuple[str, str]] = [
    ("ID", "id"),
    ("Business Name", "business_name"),
    ("Category", "category"),
    ("Rating", "rating"),
    ("Reviews", "review_count"),
    ("Business Status", "business_status"),
    ("Doctor Name", "doctor_name"),
    ("Phone", "phone"),
    ("Alternate Phone", "alternate_phone"),
    ("Email", "email"),
    ("Address", "address"),
    ("Area", "area"),
    ("City", "city"),
    ("State", "state"),
    ("Postal Code", "postal_code"),
    ("Website", "website"),
    ("Google Maps URL", "google_maps_url"),
    ("Latitude", "latitude"),
    ("Longitude", "longitude"),
    ("Source", "source"),
    ("Source ID", "source_id"),
    ("Date Found", "date_found"),
    ("Last Updated", "last_updated"),
    ("Status", "status"),
    ("Notes", "notes"),
    ("Short Info", "short_info"),
    ("Info Source", "short_info_source"),
    ("Call Status", "call_status"),
    ("Will Speak Further", "will_speak_further"),
    ("Meeting Date", "meeting_date"),
    ("Meeting Place", "meeting_place"),
    ("Follow Up", "follow_up"),
    ("Follow Up Date", "follow_up_date"),
    ("Sources", "sources"),
]

COLUMNS = [header for header, _ in _LAYOUT]
_FIELDS = [field for _, field in _LAYOUT]

_WIDTHS = {
    "Business Name": 32,
    "Doctor Name": 24,
    "Phone": 16,
    "Alternate Phone": 16,
    "Email": 28,
    "Address": 44,
    "Area": 18,
    "City": 16,
    "Website": 32,
    "Google Maps URL": 34,
    "Notes": 40,
    "Sources": 34,
    "ID": 12,
    "Rating": 9,
    "Reviews": 9,
    "Business Status": 17,
    "Short Info": 52,
    "Info Source": 13,
    "Call Status": 14,
    "Will Speak Further": 17,
    "Meeting Date": 13,
    "Meeting Place": 26,
    "Follow Up": 11,
    "Follow Up Date": 15,
    "Source ID": 18,
}

_HYPERLINK_COLUMNS = {"Website", "Google Maps URL"}



def _write_header(ws: Worksheet) -> None:
    for col, header in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)


def _apply_widths(ws: Worksheet) -> None:
    for col, header in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col)].width = _WIDTHS.get(header, 14)


def _write_row(ws: Worksheet, row: int, business: Business) -> None:
    for col, (header, field) in enumerate(_LAYOUT, start=1):
        value = getattr(business, field)

        if field == "follow_up":
            # Blank, not "False" - an absence should read as one.
            value = "Yes" if value else None

        if field == "sources":
            # dict -> readable "field: url" lines, blank when empty. Never "{}".
            value = "\n".join(f"{k}: {v}" for k, v in sorted(value.items())) or None

        cell = ws.cell(row=row, column=col)
        # None writes as a genuinely empty cell - never the string "None".
        cell.value = value

        if value and header in _HYPERLINK_COLUMNS:
            cell.hyperlink = value
            cell.style = "Hyperlink"

        if header in ("Address", "Notes", "Sources", "Short Info"):
            cell.alignment = Alignment(wrap_text=True, vertical="top")


def build(businesses: list[Business]) -> bytes:
    """Render a workbook in memory and return its bytes."""
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET
    _write_header(ws)
    _apply_widths(ws)

    for row, business in enumerate(businesses, start=2):
        _write_row(ws, row, business)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
