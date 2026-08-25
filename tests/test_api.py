import pytest
from fastapi.testclient import TestClient

from app.models import Business


class _FakeProvider:
    name = "fake"

    def search(self, query, location, limit, radius_m=None):
        return [
            Business(business_name="ABC Dental", phone="+919876543210", area=location),
            Business(business_name="Smile Care", phone="+919999999999", area=location),
        ]


@pytest.fixture
def client(dsn, monkeypatch):
    import importlib

    monkeypatch.setenv("DATABASE_URL", dsn)
    monkeypatch.delenv("APP_PASSWORD", raising=False)

    import app.main as main

    importlib.reload(main)
    monkeypatch.setattr("app.research.get_provider", lambda name=None: _FakeProvider())
    return TestClient(main.app)


def _found(client, workbook="a.xlsx"):
    client.post("/workbooks", json={"path": workbook, "kind": "workbook"})
    return client.post(
        "/search",
        json={"business_type": "dental clinic", "location": "Chromepet",
              "workbook": workbook},
    ).json()["businesses"]


def _saved(client, workbook="a.xlsx"):
    rows = _found(client, workbook)
    client.post("/businesses/save", json={"workbook": workbook, "businesses": rows})
    return client.get("/businesses", params={"workbook": workbook}).json()["businesses"]


# --- health & UI -----------------------------------------------------------


def test_health_reports_provider_and_key_status(client):
    body = client.get("/health").json()
    assert body["status"] == "ok"
    assert "provider" in body and "gemini_key_configured" in body and "model" in body


def test_ui_is_served_at_root(client):
    r = client.get("/")
    assert r.status_code == 200 and "text/html" in r.headers["content-type"]


# --- statelessness ---------------------------------------------------------


def test_search_stores_nothing(client):
    rows = _found(client)
    assert len(rows) >= 1
    assert client.get("/businesses", params={"workbook": "a.xlsx"}).json()["count"] == 0


def test_save_takes_the_rows_from_the_request_body(client):
    rows = _found(client)
    r = client.post("/businesses/save", json={"workbook": "a.xlsx", "businesses": rows})
    assert r.json()["new"] == len(rows)
    assert client.get("/businesses", params={"workbook": "a.xlsx"}).json()["count"] == len(rows)


def test_save_with_an_empty_body_is_a_clean_no_op(client):
    r = client.post("/businesses/save", json={"workbook": "a.xlsx", "businesses": []})
    assert r.status_code == 200
    assert r.json() == {"new": 0, "updated": 0, "existing": 0, "review": 0}


def test_saving_the_same_rows_twice_does_not_duplicate(client):
    rows = _found(client)
    client.post("/businesses/save", json={"workbook": "a.xlsx", "businesses": rows})
    first = client.get("/businesses", params={"workbook": "a.xlsx"}).json()["count"]
    client.post("/businesses/save", json={"workbook": "a.xlsx", "businesses": rows})
    assert client.get("/businesses", params={"workbook": "a.xlsx"}).json()["count"] == first


def test_save_targets_the_named_workbook(client):
    rows = _found(client, "a.xlsx")
    client.post("/businesses/save", json={"workbook": "a.xlsx", "businesses": rows})
    assert client.get("/businesses", params={"workbook": "a.xlsx"}).json()["count"] >= 1
    assert client.get("/businesses", params={"workbook": "b.xlsx"}).json()["count"] == 0


# --- enrichment ------------------------------------------------------------


def test_enrich_handles_one_business(client, monkeypatch):
    async def fake(business, http):
        return business.model_copy(update={"doctor_name": "Dr. Priya Kumar"})

    monkeypatch.setattr("app.main.enrich", fake)
    r = client.post(
        "/enrich",
        json={"business": {"business_name": "ABC", "website": "https://x.com"}},
    )
    assert r.status_code == 200
    assert r.json()["business"]["doctor_name"] == "Dr. Priya Kumar"
    assert r.json()["enriched"] is True


def test_enrich_without_a_website_returns_the_business_unchanged(client):
    r = client.post("/enrich", json={"business": {"business_name": "ABC"}})
    assert r.status_code == 200
    assert r.json()["business"]["business_name"] == "ABC"
    assert r.json()["enriched"] is False


def test_enrich_never_fails_a_row_over_a_bad_site(client, monkeypatch):
    async def boom(business, http):
        raise RuntimeError("connection reset")

    monkeypatch.setattr("app.main.enrich", boom)
    r = client.post(
        "/enrich",
        json={"business": {"business_name": "ABC", "website": "https://x.com"}},
    )
    assert r.status_code == 200 and r.json()["enriched"] is False


def test_enrich_tolerates_a_tag_field_from_the_client(client, monkeypatch):
    async def fake(business, http):
        return business

    monkeypatch.setattr("app.main.enrich", fake)
    r = client.post(
        "/enrich",
        json={"business": {"business_name": "ABC", "website": "https://x.com",
                           "tag": "new"}},
    )
    assert r.status_code == 200


# --- workbooks -------------------------------------------------------------


def test_workbook_tree_endpoint(client):
    assert client.get("/workbooks").json()["type"] == "folder"


def test_create_workbook_then_it_appears_in_the_tree(client):
    assert client.post(
        "/workbooks", json={"path": "dental/chennai.xlsx", "kind": "workbook"}
    ).status_code == 200
    assert "dental" in [c["name"] for c in client.get("/workbooks").json()["children"]]


def test_traversal_path_is_rejected_with_400(client):
    assert client.post(
        "/workbooks", json={"path": "../../evil.xlsx", "kind": "workbook"}
    ).status_code == 400


def test_download_rejects_traversal(client):
    assert client.get(
        "/workbooks/download", params={"path": "../../etc/passwd"}
    ).status_code == 400


def test_download_streams_a_generated_workbook(client):
    _saved(client)
    r = client.get("/workbooks/download", params={"path": "a.xlsx"})
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers["content-type"]
    assert r.content[:2] == b"PK", "a real xlsx is a zip"
    assert "a.xlsx" in r.headers["content-disposition"]


def test_rename_carries_the_rows(client):
    _saved(client)
    before = client.get("/businesses", params={"workbook": "a.xlsx"}).json()["count"]
    assert before > 0
    assert client.patch("/workbooks", json={"src": "a.xlsx", "dst": "b.xlsx"}).status_code == 200
    assert client.get("/businesses", params={"workbook": "b.xlsx"}).json()["count"] == before
    assert client.get("/businesses", params={"workbook": "a.xlsx"}).json()["count"] == 0


def test_delete_workbook_hides_its_rows(client):
    _saved(client)
    r = client.request("DELETE", "/workbooks", json={"path": "a.xlsx", "kind": "workbook"})
    assert r.status_code == 200 and r.json()["rows_removed"] >= 1
    assert client.get("/businesses", params={"workbook": "a.xlsx"}).json()["count"] == 0


def test_delete_non_empty_folder_returns_400(client):
    client.post("/workbooks", json={"path": "dental/x.xlsx", "kind": "workbook"})
    client.post("/workbooks", json={"path": "dental", "kind": "folder"})
    r = client.request("DELETE", "/workbooks", json={"path": "dental", "kind": "folder"})
    assert r.status_code == 400 and "not empty" in r.json()["detail"]


def test_open_returns_saved_rows(client):
    _saved(client)
    r = client.post("/workbooks/open", json={"workbook": "a.xlsx"})
    assert r.status_code == 200 and r.json()["count"] >= 1
    assert r.json()["adopted"] == 0


# --- editing ---------------------------------------------------------------


def test_patch_updates_a_cell(client):
    rows = _saved(client)
    r = client.patch(
        f"/businesses/{rows[0]['id']}",
        json={"changes": {"doctor_name": "Dr. Priya Kumar"}, "workbook": "a.xlsx"},
    )
    assert r.status_code == 200
    assert r.json()["business"]["doctor_name"] == "Dr. Priya Kumar"


def test_patch_appears_in_the_downloaded_workbook(client):
    from io import BytesIO

    from openpyxl import load_workbook

    from app.excel import COLUMNS

    rows = _saved(client)
    client.patch(
        f"/businesses/{rows[0]['id']}",
        json={"changes": {"doctor_name": "Dr. Priya Kumar"}, "workbook": "a.xlsx"},
    )
    data = client.get("/workbooks/download", params={"path": "a.xlsx"}).content
    ws = load_workbook(BytesIO(data))["Businesses"]
    col = COLUMNS.index("Doctor Name") + 1
    values = [ws.cell(row=i, column=col).value for i in range(2, ws.max_row + 1)]
    assert "Dr. Priya Kumar" in values


def test_patch_normalizes_a_phone(client):
    rows = _saved(client)
    r = client.patch(
        f"/businesses/{rows[0]['id']}",
        json={"changes": {"phone": "98765 43210"}, "workbook": "a.xlsx"},
    )
    assert r.json()["business"]["phone"] == "+919876543210"


def test_patch_duplicate_id_returns_400(client):
    rows = _saved(client)
    assert len(rows) >= 2
    r = client.patch(
        f"/businesses/{rows[1]['id']}",
        json={"changes": {"id": rows[0]["id"]}, "workbook": "a.xlsx"},
    )
    assert r.status_code == 400 and "already used" in r.json()["detail"]


def test_patch_bad_latitude_returns_400(client):
    rows = _saved(client)
    r = client.patch(
        f"/businesses/{rows[0]['id']}",
        json={"changes": {"latitude": "abc"}, "workbook": "a.xlsx"},
    )
    assert r.status_code == 400


def test_patch_bad_call_status_returns_400(client):
    rows = _saved(client)
    r = client.patch(
        f"/businesses/{rows[0]['id']}",
        json={"changes": {"call_status": "nonsense"}, "workbook": "a.xlsx"},
    )
    assert r.status_code == 400


def test_delete_row_endpoint(client):
    rows = _saved(client)
    before = len(rows)
    assert client.request(
        "DELETE", f"/businesses/{rows[0]['id']}", json={"workbook": "a.xlsx"}
    ).status_code == 200
    assert client.get("/businesses", params={"workbook": "a.xlsx"}).json()["count"] == before - 1


def test_add_blank_row_endpoint(client):
    _saved(client)
    r = client.post("/businesses", json={"workbook": "a.xlsx"})
    assert r.status_code == 200 and r.json()["business"]["id"]
    assert r.json()["business"]["workbook"] == "a.xlsx"


# --- commands --------------------------------------------------------------


def test_command_without_api_key_returns_400_with_actionable_message(client, monkeypatch):
    monkeypatch.delenv("GEMINI_API_KEY", raising=False)
    monkeypatch.delenv("GOOGLE_API_KEY", raising=False)
    r = client.post("/command", json={"command": "Find dental clinics in Chromepet"})
    assert r.status_code == 400 and "GEMINI_API_KEY" in r.json()["detail"]


def test_rate_limited_returns_429_not_500(client, monkeypatch):
    from app.parser import RateLimited

    monkeypatch.setattr(
        "app.main.parse_command",
        lambda text: (_ for _ in ()).throw(RateLimited("5 requests per minute")),
    )
    r = client.post("/command", json={"command": "Find dental clinics"})
    assert r.status_code == 429


def test_parser_error_returns_502_with_reason(client, monkeypatch):
    from app.parser import ParserError

    monkeypatch.setattr(
        "app.main.parse_command",
        lambda text: (_ for _ in ()).throw(ParserError("API key not valid")),
    )
    r = client.post("/command", json={"command": "Find dental clinics"})
    assert r.status_code == 502 and "API key not valid" in r.json()["detail"]


def test_search_notices_when_it_returns_fewer_than_requested(client):
    client.post("/workbooks", json={"path": "a.xlsx", "kind": "workbook"})
    body = client.post(
        "/search",
        json={"business_type": "dental clinic", "location": "Chromepet",
              "quantity": 150, "workbook": "a.xlsx"},
    ).json()
    assert body["count"] < 150 and body["notice"] and "150" in body["notice"]


def test_no_notice_when_the_request_is_satisfied(client):
    client.post("/workbooks", json={"path": "a.xlsx", "kind": "workbook"})
    body = client.post(
        "/search",
        json={"business_type": "dental clinic", "location": "Chromepet",
              "quantity": 2, "workbook": "a.xlsx"},
    ).json()
    assert body["notice"] is None
