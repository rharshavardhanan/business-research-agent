from io import BytesIO

from openpyxl import load_workbook

from app.excel import COLUMNS, build
from app.models import Business
from app.store import DEFAULT_WORKBOOK


def B(**kw):
    kw.setdefault("business_name", "X")
    return Business(**kw)


# --- store: dedupe and workbook scoping ------------------------------------


def test_upsert_new_then_existing(store):
    tagged = store.upsert_many([B(business_name="ABC Dental", phone="+919876543210")], "b.xlsx")
    assert [t for _, t in tagged] == ["new"]
    tagged = store.upsert_many([B(business_name="ABC Dental", phone="+919876543210")], "b.xlsx")
    assert [t for _, t in tagged] == ["existing"]
    assert len(store.all("b.xlsx")) == 1


def test_upsert_updates_when_new_info_arrives(store):
    store.upsert_many([B(business_name="ABC Dental", phone="+919876543210")], "b.xlsx")
    tagged = store.upsert_many(
        [B(business_name="ABC Dental", phone="+919876543210",
           website="https://abcdental.com")], "b.xlsx")
    assert [t for _, t in tagged] == ["updated"]
    rows = store.all("b.xlsx")
    assert len(rows) == 1 and rows[0].website == "https://abcdental.com"


def test_duplicates_within_one_batch_collapse(store):
    tagged = store.upsert_many(
        [B(business_name="ABC Dental", phone="+919876543210"),
         B(business_name="ABC Dental Clinic", phone="098765 43210")], "b.xlsx")
    assert [t for _, t in tagged] == ["new", "existing"]
    assert len(store.all("b.xlsx")) == 1


def test_same_business_in_two_workbooks_is_two_rows(store):
    b = B(business_name="ABC Dental", phone="+919876543210")
    assert [t for _, t in store.upsert_many([b], "a.xlsx")] == ["new"]
    assert [t for _, t in store.upsert_many([b], "b.xlsx")] == ["new"]
    assert len(store.all()) == 2
    assert len(store.all("a.xlsx")) == 1


def test_same_business_twice_in_one_workbook_is_one_row(store):
    b = B(business_name="ABC Dental", phone="+919876543210")
    store.upsert_many([b], "a.xlsx")
    assert [t for _, t in store.upsert_many([b], "a.xlsx")] == ["existing"]
    assert len(store.all("a.xlsx")) == 1


def test_filter_is_scoped_to_one_workbook(store):
    store.upsert_many([B(business_name="No Site A", phone="+919876543210")], "a.xlsx")
    store.upsert_many([B(business_name="No Site B", phone="+919876543211")], "b.xlsx")
    assert [x.business_name for x in store.filter("without_website", "a.xlsx")] == ["No Site A"]


def test_filter_without_website(store):
    store.upsert_many(
        [B(business_name="Has Site", phone="+919876543210", website="https://a.com"),
         B(business_name="No Site", phone="+919876543211")], "b.xlsx")
    assert [b.business_name for b in store.filter("without_website", "b.xlsx")] == ["No Site"]


def test_move_rows_follows_a_renamed_workbook(store):
    store.upsert_many([B(business_name="ABC")], "a.xlsx")
    store.ensure_workbook("dental/b.xlsx")
    assert store.move_rows("a.xlsx", "dental/b.xlsx") == 1
    assert len(store.all("a.xlsx")) == 0
    assert len(store.all("dental/b.xlsx")) == 1


def test_delete_rows_removes_only_that_workbook(store):
    store.upsert_many([B(business_name="A")], "a.xlsx")
    store.upsert_many([B(business_name="B")], "b.xlsx")
    assert store.delete_rows("a.xlsx") == 1
    assert len(store.all()) == 1


def test_dedupe_existing_is_scoped_to_one_workbook(store):
    b = B(business_name="ABC Dental", phone="+919876543210")
    store.upsert_many([b], "a.xlsx")
    store.upsert_many([b], "b.xlsx")
    assert store.dedupe_existing("a.xlsx") == 0, "cross-workbook rows are not duplicates"
    assert len(store.all()) == 2


def test_sources_json_round_trips(store):
    store.upsert_many([B(business_name="ABC",
                         sources={"doctor_name": "https://a.com/about"})], "b.xlsx")
    assert store.all("b.xlsx")[0].sources == {"doctor_name": "https://a.com/about"}


def test_rating_round_trips_through_the_store(store):
    store.upsert_many([B(business_name="ABC", rating=4.9, review_count=568,
                         business_status="OPERATIONAL")], "b.xlsx")
    row = store.all("b.xlsx")[0]
    assert row.rating == 4.9 and row.review_count == 568
    assert row.business_status == "OPERATIONAL"


def test_outreach_survives_a_re_search(store):
    """A re-search must never wipe the user's own observations."""
    store.upsert_many([B(business_name="ABC", phone="+919876543210")], "b.xlsx")
    row = store.all("b.xlsx")[0]
    store.update_fields(row.id, {"call_status": "Picked up", "will_speak_further": "Yes"})
    store.upsert_many([B(business_name="ABC Dental", phone="+919876543210")], "b.xlsx")
    after = store.all("b.xlsx")[0]
    assert after.call_status == "Picked up" and after.will_speak_further == "Yes"


def test_default_workbook_constant_is_used(store):
    store.upsert_many([B(business_name="ABC")])
    assert store.all(DEFAULT_WORKBOOK)[0].business_name == "ABC"


# --- excel: built in memory ------------------------------------------------


def _sheet(businesses):
    return load_workbook(BytesIO(build(businesses)))["Businesses"]


def test_build_returns_a_readable_workbook():
    ws = _sheet([B(business_name="ABC Dental", id="i", rating=4.9)])
    assert [c.value for c in ws[1]] == COLUMNS
    assert ws.cell(row=2, column=COLUMNS.index("Business Name") + 1).value == "ABC Dental"
    assert ws.cell(row=2, column=COLUMNS.index("Rating") + 1).value == 4.9


def test_build_output_is_a_real_xlsx():
    assert build([B(business_name="ABC", id="i")])[:2] == b"PK"


def test_build_keeps_the_formatting():
    ws = _sheet([B(business_name="ABC", id="i")])
    assert ws.freeze_panes == "A2"
    assert ws.auto_filter.ref is not None


def test_build_hyperlinks_the_website():
    ws = _sheet([B(business_name="A", id="i", website="https://x.com")])
    assert ws.cell(row=2, column=COLUMNS.index("Website") + 1).hyperlink is not None


def test_build_renders_follow_up_as_yes_or_blank():
    ws = _sheet([B(business_name="A", id="1", follow_up=True),
                 B(business_name="B", id="2", follow_up=False)])
    col = COLUMNS.index("Follow Up") + 1
    assert ws.cell(row=2, column=col).value == "Yes"
    assert ws.cell(row=3, column=col).value in (None, "")


def test_build_of_an_empty_list_is_still_a_valid_workbook():
    assert _sheet([]).max_row == 1


def test_null_fields_render_blank_not_the_string_none():
    ws = _sheet([B(business_name="ABC", id="i")])
    assert ws.cell(row=2, column=COLUMNS.index("Doctor Name") + 1).value in (None, "")


def test_outreach_columns_in_the_workbook():
    ws = _sheet([B(business_name="ABC", id="i", call_status="Picked up",
                   will_speak_further="Yes", meeting_date="2026-09-01",
                   meeting_place="Clinic", follow_up=True, follow_up_date="2026-09-05")])
    for name in ("Call Status", "Will Speak Further", "Meeting Date",
                 "Meeting Place", "Follow Up", "Follow Up Date"):
        assert name in COLUMNS
    assert ws.cell(row=2, column=COLUMNS.index("Call Status") + 1).value == "Picked up"


def test_short_info_columns_in_the_workbook():
    ws = _sheet([B(business_name="ABC", id="i", short_info="A dental clinic in Chennai.",
                   short_info_source="Website")])
    assert ws.cell(row=2, column=COLUMNS.index("Short Info") + 1).value == \
        "A dental clinic in Chennai."
    assert ws.cell(row=2, column=COLUMNS.index("Info Source") + 1).value == "Website"


def test_rating_columns_appear_in_the_workbook():
    ws = _sheet([B(business_name="ABC", id="i", rating=4.9, review_count=568)])
    assert ws.cell(row=2, column=COLUMNS.index("Reviews") + 1).value == 568


def test_build_reflects_the_store(store):
    store.upsert_many([B(business_name="From Store", phone="+919876543210")], "b.xlsx")
    ws = _sheet(store.all("b.xlsx"))
    assert ws.cell(row=2, column=COLUMNS.index("Business Name") + 1).value == "From Store"
