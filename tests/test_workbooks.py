import pytest
from openpyxl import load_workbook

from app import workbooks
from app.excel import COLUMNS
from app.workbooks import (
    InvalidPath,
    create_folder,
    create_workbook,
    delete_folder,
    delete_workbook,
    rename_or_move,
    resolve_path,
    tree,
)


@pytest.fixture(autouse=True)
def data_dir(tmp_path, monkeypatch):
    monkeypatch.setattr(workbooks, "DATA_DIR", tmp_path)
    return tmp_path


# --- path safety (trust boundary) -----------------------------------------


@pytest.mark.parametrize(
    "evil",
    ["../../etc/passwd", "/etc/passwd", "a/../../b.xlsx", "../leads.db", "..", "", "   "],
)
def test_traversal_and_absolute_paths_are_rejected(evil):
    with pytest.raises(InvalidPath):
        resolve_path(evil)


def test_non_xlsx_workbook_path_is_rejected():
    with pytest.raises(InvalidPath):
        resolve_path("notes.txt", require_xlsx=True)


def test_the_database_is_not_a_valid_target():
    with pytest.raises(InvalidPath):
        resolve_path("leads.db")


def test_trash_is_not_a_valid_target():
    with pytest.raises(InvalidPath):
        resolve_path(".trash/old.xlsx")


def test_a_normal_path_resolves_inside_data_dir(data_dir):
    got = resolve_path("dental/chennai.xlsx")
    assert got == (data_dir / "dental" / "chennai.xlsx").resolve()
    assert data_dir.resolve() in got.parents


def test_symlink_escaping_data_dir_is_rejected(data_dir, tmp_path):
    outside = tmp_path.parent / "outside_secret"
    outside.mkdir(exist_ok=True)
    (data_dir / "escape").symlink_to(outside, target_is_directory=True)
    with pytest.raises(InvalidPath):
        resolve_path("escape/secret.xlsx")


def test_tree_lists_workbooks_and_folders(data_dir):
    (data_dir / "businesses.xlsx").write_bytes(b"x")
    (data_dir / "dental").mkdir()
    (data_dir / "dental" / "chennai.xlsx").write_bytes(b"x")
    (data_dir / "leads.db").write_bytes(b"x")
    (data_dir / ".trash").mkdir()
    (data_dir / ".trash" / "gone.xlsx").write_bytes(b"x")

    t = tree()
    names = [c["name"] for c in t["children"]]
    assert "businesses.xlsx" in names
    assert "dental" in names
    assert "leads.db" not in names, "the database is not a workbook"
    assert ".trash" not in names, "trash is never listed"

    dental = next(c for c in t["children"] if c["name"] == "dental")
    assert dental["type"] == "folder"
    assert [c["path"] for c in dental["children"]] == ["dental/chennai.xlsx"]


# --- operations -----------------------------------------------------------


def test_create_workbook_writes_a_formatted_empty_sheet(data_dir):
    rel = create_workbook("dental/chennai.xlsx")
    assert rel == "dental/chennai.xlsx"
    ws = load_workbook(data_dir / "dental" / "chennai.xlsx")["Businesses"]
    assert [c.value for c in ws[1]] == COLUMNS
    assert ws.max_row == 1


def test_create_workbook_refuses_to_clobber(data_dir):
    create_workbook("a.xlsx")
    with pytest.raises(InvalidPath, match="already exists"):
        create_workbook("a.xlsx")


def test_create_workbook_requires_xlsx():
    with pytest.raises(InvalidPath):
        create_workbook("notes.txt")


def test_create_folder(data_dir):
    assert create_folder("dental/south") == "dental/south"
    assert (data_dir / "dental" / "south").is_dir()


def test_rename_moves_the_file(data_dir):
    create_workbook("a.xlsx")
    assert rename_or_move("a.xlsx", "dental/b.xlsx") == "dental/b.xlsx"
    assert not (data_dir / "a.xlsx").exists()
    assert (data_dir / "dental" / "b.xlsx").exists()


def test_rename_refuses_to_overwrite(data_dir):
    create_workbook("a.xlsx")
    create_workbook("b.xlsx")
    with pytest.raises(InvalidPath, match="already exists"):
        rename_or_move("a.xlsx", "b.xlsx")


def test_delete_workbook_moves_to_trash_never_unlinks(data_dir):
    create_workbook("a.xlsx")
    trashed = delete_workbook("a.xlsx")
    assert not (data_dir / "a.xlsx").exists()
    assert (data_dir / ".trash" / trashed).exists(), "the spreadsheet is the backup"


def test_delete_empty_folder(data_dir):
    create_folder("empty")
    delete_folder("empty")
    assert not (data_dir / "empty").exists()


def test_delete_non_empty_folder_is_refused(data_dir):
    create_workbook("dental/chennai.xlsx")
    with pytest.raises(InvalidPath, match="not empty"):
        delete_folder("dental")
    assert (data_dir / "dental" / "chennai.xlsx").exists()
